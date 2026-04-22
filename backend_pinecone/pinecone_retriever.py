from __future__ import annotations

"""Pinecone-backed embeddings retriever and data loaders.

This module provides:
- Chunking (recursive splitter with overlap)
- Loading policy docs from JSON and playbook entries from XLSX
- Upserting/querying vectors in Pinecone (per-namespace)

The embedding model runs locally (`sentence-transformers`); Pinecone stores vectors + metadata.
"""

import hashlib
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

import openpyxl
from sentence_transformers import SentenceTransformer


# ---------------------------- Small helpers ----------------------------

def _norm_header(value: object) -> str:
    s = str(value or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("_", " ")
    return s


def _stable_id(*parts: str) -> str:
    h = hashlib.sha1()
    for p in parts:
        h.update((p or "").encode("utf-8"))
        h.update(b"\x1f")
    return h.hexdigest()


# ---------------------------- Chunking ----------------------------
def chunk_text(text: str, *, max_chars: int = 900, overlap: int = 160) -> list[str]:
    """Split text into overlapping chunks using recursive, structure-aware splitting."""

    raw = (text or "")
    raw = raw.replace("\r\n", "\n").replace("\r", "\n")
    raw = re.sub(r"[\t\f\v]+", " ", raw)
    raw = re.sub(r" +", " ", raw)
    t = raw.strip()
    if not t:
        return []
    if len(t) <= max_chars:
        return [t]

    # Ordered from “coarsest” boundary to “finest”.
    # `joiner` is used when we merge small pieces back into max-sized chunks.
    separators: list[tuple[str, str]] = [
        (r"\n{2,}", "\n\n"),
        (r"\n", "\n"),
        (r"(?<=[.!?])\s+", " "),
        (r";\s+", "; "),
        (r",\s+", ", "),
        (r"\s+", " "),
    ]

    def _char_chunks(s: str) -> list[str]:
        # Last-resort fallback when no separators produce multiple parts.
        out: list[str] = []
        start = 0
        while start < len(s):
            end = min(len(s), start + max_chars)
            chunk = s[start:end]
            if end < len(s) and " " in chunk:
                chunk = chunk.rsplit(" ", 1)[0]
                end = start + len(chunk)
            chunk = chunk.strip()
            if chunk:
                out.append(chunk)
            if end >= len(s):
                break
            start = max(0, end - max(0, overlap))
        return out

    def _merge(parts: list[str], joiner: str) -> list[str]:
        # Greedily pack parts into chunks up to `max_chars`.
        merged: list[str] = []
        buf: list[str] = []
        buf_len = 0
        for p in parts:
            p = p.strip()
            if not p:
                continue
            cand_len = len(p) if not buf else buf_len + len(joiner) + len(p)
            if cand_len <= max_chars:
                if buf:
                    buf_len += len(joiner) + len(p)
                else:
                    buf_len = len(p)
                buf.append(p)
                continue

            if buf:
                merged.append(joiner.join(buf).strip())
            buf = [p]
            buf_len = len(p)

        if buf:
            merged.append(joiner.join(buf).strip())
        return [m for m in merged if m]

    def _split_recursive(s: str, seps: list[tuple[str, str]]) -> list[str]:
        # Split by the first separator that yields >1 part, then recursively
        # re-split any overlong parts using the remaining “finer” separators.
        s = s.strip()
        if not s:
            return []
        if len(s) <= max_chars:
            return [s]
        if not seps:
            return _char_chunks(s)

        for idx, (pattern, joiner) in enumerate(seps):
            parts = [p.strip() for p in re.split(pattern, s) if p and p.strip()]
            if len(parts) <= 1:
                continue

            refined: list[str] = []
            for p in parts:
                if len(p) > max_chars:
                    refined.extend(_split_recursive(p, seps[idx + 1 :]))
                else:
                    refined.append(p)

            merged = _merge(refined, joiner)
            if merged and len(merged) > 1:
                return merged

        return _char_chunks(s)

    base_chunks = _split_recursive(t, separators)

    if overlap <= 0 or len(base_chunks) <= 1:
        return base_chunks

    out: list[str] = [base_chunks[0]]
    for i in range(1, len(base_chunks)):
        prev = base_chunks[i - 1]
        prefix = prev[-overlap:] if overlap < len(prev) else prev
        if " " in prefix:
            prefix = prefix.split(" ", 1)[-1]
        combined = (prefix + " " + base_chunks[i]).strip() if prefix else base_chunks[i]
        out.append(combined)

    return out


# ---------------------------- Data structures ----------------------------
@dataclass(frozen=True)
class SourceDoc:
    title: str
    content: str


@dataclass(frozen=True)
class RetrievedDoc:
    title: str
    content: str
    score: float


# ---------------------------- Pinecone connectivity ----------------------------
def _connect_pinecone(*, api_key: str, index_name: str, host: Optional[str] = None):
    """Connect to Pinecone across SDK versions.

    Supports:
    - New SDK: from pinecone import Pinecone
    - Old SDK: import pinecone; pinecone.init(...)
    """

    def _norm_host(h: Optional[str]) -> Optional[str]:
        # Accept either a raw host or a full URL; normalize to what Pinecone SDKs expect.
        if not h:
            return None
        hh = str(h).strip()
        if not hh:
            return None
        # Pinecone SDKs typically expect just the host (no scheme).
        hh = re.sub(r"^https?://", "", hh, flags=re.IGNORECASE)
        hh = hh.rstrip("/")
        return hh or None

    host = _norm_host(host)

    try:
        from pinecone import Pinecone  # type: ignore

        pc = Pinecone(api_key=api_key)
        if host:
            return pc.Index(host=host)

        # Most common: Index(name)
        try:
            return pc.Index(index_name)
        except TypeError:
            # Some variants: Index(name=...)
            return pc.Index(name=index_name)

    except Exception:
        import pinecone  # type: ignore

        # Old client requires environment/region for init.
        env = (
            Path(".")  # placeholder for mypy
            and (
                __import__("os").environ.get("PINECONE_ENVIRONMENT")
                or __import__("os").environ.get("PINECONE_REGION")
                or __import__("os").environ.get("PINECONE_ENV")
            )
        )
        if env:
            pinecone.init(api_key=api_key, environment=env)
        else:
            # Some older versions allow init without env if using a host-based Index.
            pinecone.init(api_key=api_key)

        if host and hasattr(pinecone, "Index"):
            try:
                return pinecone.Index(host=host)
            except TypeError:
                pass
        return pinecone.Index(index_name)


# ---------------------------- Pinecone index ----------------------------
class PineconePolicyIndex:
    """Embeddings-based retriever backed by a Pinecone index.

    Stores chunk text inside metadata under key `text`.
    """

    def __init__(
        self,
        *,
        api_key: str,
        index_name: str,
        namespace: str,
        embedding_model_name: str = "all-MiniLM-L6-v2",
        docs: Iterable[SourceDoc] = (),
        source_label: str,
        host: Optional[str] = None,
        chunk_max_chars: int = 900,
        chunk_overlap: int = 160,
        upsert_batch_size: int = 96,
    ):
        self._namespace = namespace
        self._source_label = source_label
        self._chunk_max_chars = chunk_max_chars
        self._chunk_overlap = chunk_overlap
        self._upsert_batch_size = max(1, upsert_batch_size)

        # Local embedding model used to generate vectors for upsert/query.
        self._model = SentenceTransformer(embedding_model_name)
        self._index = _connect_pinecone(api_key=api_key, index_name=index_name, host=host)

        docs_list = list(docs)
        if docs_list:
            if self._is_namespace_empty():
                self.upsert(docs_list)

    def _is_namespace_empty(self) -> bool:
        # Used to avoid re-upserting the same corpus on every startup.
        try:
            stats = self._index.describe_index_stats()  # type: ignore
            # New/old SDKs return dict-like objects.
            ns = None
            if isinstance(stats, dict):
                ns = (stats.get("namespaces") or {}).get(self._namespace)
            else:
                ns = getattr(stats, "namespaces", {}).get(self._namespace)
            if not ns:
                return True
            # Pinecone typically returns {"vector_count": N}
            vc = ns.get("vector_count") if isinstance(ns, dict) else getattr(ns, "vector_count", None)
            return int(vc or 0) == 0
        except Exception:
            return False

    @staticmethod
    def docs_from_json_file(path: str | Path) -> list[SourceDoc]:
        p = Path(path)
        data = json.loads(p.read_text(encoding="utf-8"))
        return [SourceDoc(title=item["title"], content=item["content"]) for item in data]

    @staticmethod
    def docs_from_xlsx_file(path: str | Path, title_prefix: str = "Complaint Playbook") -> list[SourceDoc]:
        """Load an .xlsx where each row is a support playbook entry."""
        p = Path(path)
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]

        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return []

        header_norm = [_norm_header(h) for h in header]
        col_index = {name: idx for idx, name in enumerate(header_norm) if name}

        def get(row: tuple[object, ...], key: str) -> str:
            idx = col_index.get(key)
            if idx is None or idx >= len(row):
                return ""
            v = row[idx]
            return str(v).strip() if v is not None else ""

        docs: list[SourceDoc] = []
        for row in rows:
            trouble = get(row, "trouble")
            category = get(row, "category")
            solution = get(row, "solution")
            alt_solution = get(row, "alternate solution")
            company_response = get(row, "company response")

            if not any([trouble, category, solution, alt_solution, company_response]):
                continue

            title_bits = [title_prefix]
            if trouble:
                title_bits.append(trouble)
            if category:
                title_bits.append(f"({category})")
            title = ": ".join([title_bits[0], " ".join(title_bits[1:]).strip()]).strip(": ")

            content_parts: list[str] = []
            if category:
                content_parts.append(f"Category: {category}")
            if solution:
                content_parts.append(f"Suggested Solution: {solution}")
            if alt_solution:
                content_parts.append(f"Alternate Solution: {alt_solution}")
            if company_response:
                content_parts.append(f"Company Response Template: {company_response}")

            docs.append(SourceDoc(title=title, content="\n".join(content_parts).strip()))

        return docs

    @classmethod
    def from_sources(
        cls,
        *,
        api_key: str,
        index_name: str,
        namespace: str,
        source_label: str,
        embedding_model_name: str = "all-MiniLM-L6-v2",
        host: Optional[str] = None,
        json_path: Optional[str | Path] = None,
        xlsx_path: Optional[str | Path] = None,
        xlsx_title_prefix: str = "Complaint Playbook",
        chunk_max_chars: int = 900,
        chunk_overlap: int = 160,
    ) -> "PineconePolicyIndex":
        docs: list[SourceDoc] = []
        if json_path is not None:
            docs.extend(cls.docs_from_json_file(json_path))
        if xlsx_path is not None:
            docs.extend(cls.docs_from_xlsx_file(xlsx_path, title_prefix=xlsx_title_prefix))

        return cls(
            api_key=api_key,
            index_name=index_name,
            namespace=namespace,
            embedding_model_name=embedding_model_name,
            docs=docs,
            source_label=source_label,
            host=host,
            chunk_max_chars=chunk_max_chars,
            chunk_overlap=chunk_overlap,
        )

    def upsert(self, docs: Iterable[SourceDoc]) -> None:
        # We upsert *chunks* rather than whole documents so retrieval can focus on
        # the most relevant passage.
        ids: list[str] = []
        texts: list[str] = []
        metas: list[dict[str, object]] = []

        for doc in docs:
            title = (doc.title or "").strip()
            content = (doc.content or "").strip()
            if not (title or content):
                continue

            pieces = chunk_text(content, max_chars=self._chunk_max_chars, overlap=self._chunk_overlap)
            if not pieces:
                continue

            # Stable IDs keep re-indexing idempotent.
            base_id = _stable_id(self._source_label, title)
            for i, chunk in enumerate(pieces):
                chunk_id = _stable_id(base_id, str(i), chunk)
                ids.append(chunk_id)
                texts.append(chunk)
                metas.append({"source": self._source_label, "title": title, "chunk": i, "text": chunk})

        if not ids:
            return

        # Embed in batches to keep memory reasonable.
        for start in range(0, len(ids), self._upsert_batch_size):
            end = min(len(ids), start + self._upsert_batch_size)
            batch_ids = ids[start:end]
            batch_texts = texts[start:end]
            batch_metas = metas[start:end]

            # `normalize_embeddings=True` makes dot-product behave like cosine similarity.
            vectors = self._model.encode(batch_texts, normalize_embeddings=True).tolist()

            # Pinecone SDKs accept either tuple payloads or dict payloads depending on version.
            tuples_payload = [(i, v, m) for i, v, m in zip(batch_ids, vectors, batch_metas)]
            dict_payload = [{"id": i, "values": v, "metadata": m} for i, v, m in zip(batch_ids, vectors, batch_metas)]

            try:
                self._index.upsert(vectors=tuples_payload, namespace=self._namespace)  # type: ignore
            except Exception:
                self._index.upsert(vectors=dict_payload, namespace=self._namespace)  # type: ignore

    def search(self, query: str, top_k: int = 3) -> list[RetrievedDoc]:
        q = (query or "").strip()
        if not q:
            return []

        # Pinecone query uses dot-product / cosine similarity depending on index configuration.
        # We store chunk text in metadata, so we can reconstruct the answer context.
        q_vec = self._model.encode([q], normalize_embeddings=True).tolist()[0]

        result = self._index.query(
            vector=q_vec,
            top_k=top_k,
            include_metadata=True,
            namespace=self._namespace,
        )

        matches = None
        if isinstance(result, dict):
            matches = result.get("matches")
        else:
            matches = getattr(result, "matches", None)

        matches = matches or []
        out: list[RetrievedDoc] = []
        for m in matches:
            if isinstance(m, dict):
                score = float(m.get("score") or 0.0)
                meta = m.get("metadata") or {}
            else:
                score = float(getattr(m, "score", 0.0) or 0.0)
                meta = getattr(m, "metadata", {}) or {}

            title = str(meta.get("title") or "").strip() if isinstance(meta, dict) else ""
            text = str(meta.get("text") or "").strip() if isinstance(meta, dict) else ""
            out.append(RetrievedDoc(title=title, content=text, score=score))

        return out
