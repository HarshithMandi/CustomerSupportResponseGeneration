from __future__ import annotations

"""ChromaDB-backed embeddings retriever and data loaders.

This module owns:
- Chunking (recursive splitter with overlap)
- Loading policy docs from JSON and playbook entries from XLSX
- Creating/upserting/searching a persistent ChromaDB collection

It is intentionally self-contained so `backend/main.py` can import a single retriever.
"""

import hashlib
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

import openpyxl


# ---------------------------- Small helpers ----------------------------

def _norm_header(value: object) -> str:
    s = str(value or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("_", " ")
    return s


def _stable_id(*parts: str) -> str:
    h = hashlib.sha1()
    for p in parts:
        h.update((p or "").encode("utf-8"))
        h.update(b"\x1f")
    return h.hexdigest()



# ---------------------------- Chunking ----------------------------
def chunk_text(text: str, *, max_chars: int = 900, overlap: int = 160) -> list[str]:
    """Split text into overlapping chunks using recursive, structure-aware splitting.

    Strategy: try to split on larger boundaries first (paragraphs/newlines), then
    sentences, then phrases, then whitespace; fall back to fixed-size char chunks.
    """

    raw = (text or "")
    # Preserve newlines for better boundaries; normalize whitespace a bit.
    raw = raw.replace("\r\n", "\n").replace("\r", "\n")
    raw = re.sub(r"[\t\f\v]+", " ", raw)
    raw = re.sub(r" +", " ", raw)
    t = raw.strip()
    if not t:
        return []
    if len(t) <= max_chars:
        return [t]

    # Ordered from “coarsest” boundary to “finest”.
    # `joiner` is used when we merge small pieces back into max-sized chunks.
    separators: list[tuple[str, str]] = [
        (r"\n{2,}", "\n\n"),
        (r"\n", "\n"),
        (r"(?<=[.!?])\s+", " "),
        (r";\s+", "; "),
        (r",\s+", ", "),
        (r"\s+", " "),
    ]

    def _char_chunks(s: str) -> list[str]:
        # Last-resort fallback when no separators produce multiple parts.
        out: list[str] = []
        start = 0
        while start < len(s):
            end = min(len(s), start + max_chars)
            chunk = s[start:end]
            if end < len(s) and " " in chunk:
                chunk = chunk.rsplit(" ", 1)[0]
                end = start + len(chunk)
            chunk = chunk.strip()
            if chunk:
                out.append(chunk)
            if end >= len(s):
                break
            start = max(0, end - max(0, overlap))
        return out

    def _merge(parts: list[str], joiner: str) -> list[str]:
        # Greedily pack parts into chunks up to `max_chars`.
        merged: list[str] = []
        buf: list[str] = []
        buf_len = 0
        for p in parts:
            p = p.strip()
            if not p:
                continue
            cand_len = len(p) if not buf else buf_len + len(joiner) + len(p)
            if cand_len <= max_chars:
                if buf:
                    buf_len += len(joiner) + len(p)
                else:
                    buf_len = len(p)
                buf.append(p)
                continue

            if buf:
                merged.append(joiner.join(buf).strip())
            buf = [p]
            buf_len = len(p)

        if buf:
            merged.append(joiner.join(buf).strip())
        return [m for m in merged if m]

    def _split_recursive(s: str, seps: list[tuple[str, str]]) -> list[str]:
        # Split by the first separator that yields >1 part, then recursively
        # re-split any overlong parts using the remaining “finer” separators.
        s = s.strip()
        if not s:
            return []
        if len(s) <= max_chars:
            return [s]
        if not seps:
            return _char_chunks(s)

        for idx, (pattern, joiner) in enumerate(seps):
            parts = [p.strip() for p in re.split(pattern, s) if p and p.strip()]
            if len(parts) <= 1:
                continue

            refined: list[str] = []
            for p in parts:
                if len(p) > max_chars:
                    refined.extend(_split_recursive(p, seps[idx + 1 :]))
                else:
                    refined.append(p)

            merged = _merge(refined, joiner)
            if merged and len(merged) > 1:
                return merged

        return _char_chunks(s)

    base_chunks = _split_recursive(t, separators)

    if overlap <= 0 or len(base_chunks) <= 1:
        return base_chunks

    # Add overlap by prefixing each chunk with the tail of the previous chunk.
    out: list[str] = [base_chunks[0]]
    for i in range(1, len(base_chunks)):
        prev = base_chunks[i - 1]
        prefix = prev[-overlap:] if overlap < len(prev) else prev
        # Avoid starting the prefix mid-word when possible.
        if " " in prefix:
            prefix = prefix.split(" ", 1)[-1]
        combined = (prefix + " " + base_chunks[i]).strip() if prefix else base_chunks[i]
        out.append(combined)

    return out


# ---------------------------- Data structures ----------------------------
@dataclass(frozen=True)
class SourceDoc:
    title: str
    content: str


@dataclass(frozen=True)
class RetrievedDoc:
    title: str
    content: str
    score: float



# ---------------------------- ChromaDB index ----------------------------
class ChromaPolicyIndex:
    """Embeddings-based retriever backed by a persistent ChromaDB collection."""

    def __init__(
        self,
        *,
        collection_name: str,
        persist_dir: str | Path,
        # Local embedding model used by Chroma's SentenceTransformer embedding function.
        embedding_model_name: str = "all-MiniLM-L6-v2",
        docs: Iterable[SourceDoc] = (),
        source_label: str,
        chunk_max_chars: int = 900,
        chunk_overlap: int = 160,
    ):
        import chromadb
        from chromadb.config import Settings
        from chromadb.utils.embedding_functions import SentenceTransformerEmbeddingFunction

        self._collection_name = collection_name
        self._persist_dir = str(persist_dir)
        self._source_label = source_label
        self._chunk_max_chars = chunk_max_chars
        self._chunk_overlap = chunk_overlap

        client = chromadb.PersistentClient(
            path=self._persist_dir,
            settings=Settings(
                anonymized_telemetry=False,
            ),
        )
        embedding_fn = SentenceTransformerEmbeddingFunction(model_name=embedding_model_name)

        # Use cosine space so we can convert distance -> similarity via (1 - distance).
        self._collection = client.get_or_create_collection(
            name=self._collection_name,
            embedding_function=embedding_fn,
            metadata={"hnsw:space": "cosine"},
        )

        docs_list = list(docs)
        if docs_list:
            try:
                if self._collection.count() == 0:
                    self.upsert(docs_list)
            except Exception:
                # If count() fails for any reason, fall back to ensuring docs exist.
                self.upsert(docs_list)

    @staticmethod
    def docs_from_json_file(path: str | Path) -> list[SourceDoc]:
        p = Path(path)
        data = json.loads(p.read_text(encoding="utf-8"))
        return [SourceDoc(title=item["title"], content=item["content"]) for item in data]

    @staticmethod
    def docs_from_xlsx_file(path: str | Path, title_prefix: str = "Complaint Playbook") -> list[SourceDoc]:
        """Load an .xlsx where each row is a support playbook entry.

        Expected headers (case-insensitive):
        - Trouble, Category, Solution, Alternate Solution, Company Response
        """

        p = Path(path)
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]

        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return []

        header_norm = [_norm_header(h) for h in header]
        col_index = {name: idx for idx, name in enumerate(header_norm) if name}

        def get(row: tuple[object, ...], key: str) -> str:
            idx = col_index.get(key)
            if idx is None or idx >= len(row):
                return ""
            v = row[idx]
            return str(v).strip() if v is not None else ""

        docs: list[SourceDoc] = []
        for row in rows:
            trouble = get(row, "trouble")
            category = get(row, "category")
            solution = get(row, "solution")
            alt_solution = get(row, "alternate solution")
            company_response = get(row, "company response")

            if not any([trouble, category, solution, alt_solution, company_response]):
                continue

            title_bits = [title_prefix]
            if trouble:
                title_bits.append(trouble)
            if category:
                title_bits.append(f"({category})")
            title = ": ".join([title_bits[0], " ".join(title_bits[1:]).strip()]).strip(": ")

            content_parts: list[str] = []
            if category:
                content_parts.append(f"Category: {category}")
            if solution:
                content_parts.append(f"Suggested Solution: {solution}")
            if alt_solution:
                content_parts.append(f"Alternate Solution: {alt_solution}")
            if company_response:
                content_parts.append(f"Company Response Template: {company_response}")

            docs.append(SourceDoc(title=title, content="\n".join(content_parts).strip()))

        return docs

    @classmethod
    def from_sources(
        cls,
        *,
        collection_name: str,
        persist_dir: str | Path,
        source_label: str,
        embedding_model_name: str = "all-MiniLM-L6-v2",
        json_path: Optional[str | Path] = None,
        xlsx_path: Optional[str | Path] = None,
        xlsx_title_prefix: str = "Complaint Playbook",
        chunk_max_chars: int = 900,
        chunk_overlap: int = 160,
    ) -> "ChromaPolicyIndex":
        docs: list[SourceDoc] = []
        if json_path is not None:
            docs.extend(cls.docs_from_json_file(json_path))
        if xlsx_path is not None:
            docs.extend(cls.docs_from_xlsx_file(xlsx_path, title_prefix=xlsx_title_prefix))

        return cls(
            collection_name=collection_name,
            persist_dir=persist_dir,
            embedding_model_name=embedding_model_name,
            docs=docs,
            source_label=source_label,
            chunk_max_chars=chunk_max_chars,
            chunk_overlap=chunk_overlap,
        )

    def upsert(self, docs: Iterable[SourceDoc]) -> None:
        # We upsert *chunks* rather than whole documents so retrieval can focus on
        # the most relevant passage.
        ids: list[str] = []
        documents: list[str] = []
        metadatas: list[dict[str, object]] = []

        for doc in docs:
            title = (doc.title or "").strip()
            content = (doc.content or "").strip()
            if not (title or content):
                continue

            pieces = chunk_text(content, max_chars=self._chunk_max_chars, overlap=self._chunk_overlap)
            if not pieces:
                continue

            # Stable IDs keep re-indexing idempotent.
            base_id = _stable_id(self._source_label, title)
            for i, chunk in enumerate(pieces):
                chunk_id = _stable_id(base_id, str(i), chunk)
                ids.append(chunk_id)
                documents.append(chunk)
                metadatas.append(
                    {
                        "source": self._source_label,
                        "title": title,
                        "chunk": i,
                    }
                )

        if not ids:
            return

        self._collection.upsert(ids=ids, documents=documents, metadatas=metadatas)

    def search(self, query: str, top_k: int = 3) -> list[RetrievedDoc]:
        q = (query or "").strip()
        if not q:
            return []

        # Chroma returns cosine *distance*; we convert to a similarity score.
        result = self._collection.query(
            query_texts=[q],
            n_results=top_k,
            include=["documents", "metadatas", "distances"],
        )

        docs = (result.get("documents") or [[]])[0]
        metas = (result.get("metadatas") or [[]])[0]
        dists = (result.get("distances") or [[]])[0]

        out: list[RetrievedDoc] = []
        for doc_text, meta, dist in zip(docs, metas, dists):
            title = ""
            if isinstance(meta, dict):
                title = str(meta.get("title") or "").strip()
            # With cosine distance: similarity ~= 1 - distance.
            try:
                score = 1.0 - float(dist)
            except Exception:
                score = 0.0
            out.append(RetrievedDoc(title=title, content=str(doc_text or "").strip(), score=score))

        return out
