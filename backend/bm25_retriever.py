from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

import openpyxl
from rank_bm25 import BM25Okapi


_WORD_RE = re.compile(r"[A-Za-z0-9]+")


def tokenize(text: str) -> list[str]:
    return _WORD_RE.findall((text or "").lower())


@dataclass(frozen=True)
class PolicyDoc:
    title: str
    content: str


@dataclass(frozen=True)
class RetrievedDoc:
    title: str
    content: str
    score: float


def _norm_header(value: object) -> str:
    s = str(value or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("_", " ")
    return s


class BM25PolicyIndex:
    def __init__(self, docs: Iterable[PolicyDoc]):
        self._docs: list[PolicyDoc] = list(docs)
        corpus_tokens = [tokenize(d.title + "\n" + d.content) for d in self._docs]
        # BM25Okapi requires a non-empty corpus; keep search() safe if docs is empty.
        self._bm25 = BM25Okapi(corpus_tokens or [[]])

    @staticmethod
    def docs_from_json_file(path: str | Path) -> list[PolicyDoc]:
        p = Path(path)
        data = json.loads(p.read_text(encoding="utf-8"))
        return [PolicyDoc(title=item["title"], content=item["content"]) for item in data]

    @staticmethod
    def docs_from_xlsx_file(path: str | Path, title_prefix: str = "Complaint Playbook") -> list[PolicyDoc]:
        """Load an .xlsx where each row is a support playbook entry.

        Expected headers (case-insensitive):
        - Trouble, Category, Solution, Alternate Solution, Company Response
        """

        p = Path(path)
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]

        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return []

        header_norm = [_norm_header(h) for h in header]
        col_index = {name: idx for idx, name in enumerate(header_norm) if name}

        def get(row: tuple[object, ...], key: str) -> str:
            idx = col_index.get(key)
            if idx is None or idx >= len(row):
                return ""
            v = row[idx]
            return str(v).strip() if v is not None else ""

        docs: list[PolicyDoc] = []
        for row in rows:
            trouble = get(row, "trouble")
            category = get(row, "category")
            solution = get(row, "solution")
            alt_solution = get(row, "alternate solution")
            company_response = get(row, "company response")

            if not any([trouble, category, solution, alt_solution, company_response]):
                continue

            title_bits = [title_prefix]
            if trouble:
                title_bits.append(trouble)
            if category:
                title_bits.append(f"({category})")
            title = ": ".join([title_bits[0], " ".join(title_bits[1:]).strip()]).strip(": ")

            content_parts: list[str] = []
            if category:
                content_parts.append(f"Category: {category}")
            if solution:
                content_parts.append(f"Suggested Solution: {solution}")
            if alt_solution:
                content_parts.append(f"Alternate Solution: {alt_solution}")
            if company_response:
                content_parts.append(f"Company Response Template: {company_response}")

            docs.append(PolicyDoc(title=title, content="\n".join(content_parts).strip()))

        return docs

    @classmethod
    def from_sources(
        cls,
        *,
        json_path: Optional[str | Path] = None,
        xlsx_path: Optional[str | Path] = None,
        xlsx_title_prefix: str = "Complaint Playbook",
    ) -> "BM25PolicyIndex":
        docs: list[PolicyDoc] = []
        if json_path is not None:
            docs.extend(cls.docs_from_json_file(json_path))
        if xlsx_path is not None:
            docs.extend(cls.docs_from_xlsx_file(xlsx_path, title_prefix=xlsx_title_prefix))
        return cls(docs)

    @classmethod
    def from_json_file(cls, path: str | Path) -> "BM25PolicyIndex":
        return cls(cls.docs_from_json_file(path))

    def search(self, query: str, top_k: int = 3) -> list[RetrievedDoc]:
        if not self._docs:
            return []

        q_tokens = tokenize(query)
        if not q_tokens:
            return []

        scores = self._bm25.get_scores(q_tokens)
        ranked = sorted(enumerate(scores), key=lambda x: x[1], reverse=True)[:top_k]
        out: list[RetrievedDoc] = []
        for idx, score in ranked:
            doc = self._docs[idx]
            out.append(RetrievedDoc(title=doc.title, content=doc.content, score=float(score)))
        return out
